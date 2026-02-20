from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session ,send_file
from models import db, User, Post, Comment, Inventory, SmartringInventory
from functools import wraps
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
from io import BytesIO
import os
import re

def read_excel_fallback(file):
    """openpyxl 실패 시 직접 XML 파싱으로 Excel 읽기"""
    import zipfile
    import xml.etree.ElementTree as ET
    
    file.seek(0)
    
    with zipfile.ZipFile(file, 'r') as z:
        # 공유 문자열 읽기
        shared_strings = []
        if 'xl/sharedStrings.xml' in z.namelist():
            with z.open('xl/sharedStrings.xml') as f:
                tree = ET.parse(f)
                root = tree.getroot()
                ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                for si in root.findall('.//main:si', ns):
                    t = si.find('.//main:t', ns)
                    shared_strings.append(t.text if t is not None and t.text else '')
        
        # 시트 데이터 읽기
        with z.open('xl/worksheets/sheet1.xml') as f:
            tree = ET.parse(f)
            root = tree.getroot()
            ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            
            rows_data = []
            for row in root.findall('.//main:row', ns):
                row_data = {}
                for cell in row.findall('main:c', ns):
                    cell_ref = cell.get('r')
                    col = ''.join(filter(str.isalpha, cell_ref))
                    cell_type = cell.get('t')
                    v = cell.find('main:v', ns)
                    
                    if v is not None and v.text:
                        if cell_type == 's':
                            value = shared_strings[int(v.text)]
                        else:
                            value = v.text
                    else:
                        value = ''
                    row_data[col] = value
                
                if row_data:
                    rows_data.append(row_data)
    
    if not rows_data:
        return pd.DataFrame()
    
    # 첫 번째 행을 헤더로 사용
    headers = rows_data[0]
    col_mapping = {k: v for k, v in headers.items()}
    
    data_rows = []
    for row in rows_data[1:]:
        new_row = {}
        for col, header in col_mapping.items():
            new_row[header] = row.get(col, '')
        data_rows.append(new_row)
    
    return pd.DataFrame(data_rows)

app = Flask(__name__)

UPLOAD_FOLDER = 'static/uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

basdir = os.path.abspath(os.path.dirname(__file__))
dbfile = os.path.join(basdir, 'db.sqlite')

app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + dbfile
app.config['SQLALCHEMY_COMMIT_ON_TEARDOWN'] = True
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'jqiowejrojzxcovnklqnweiorjqwoijroi'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

db.init_app(app)
   
@app.before_request
def make_session_permanent():
    session.permanent = True
    if request.endpoint == 'register':
        pass
    
def is_admin():
    if 'userid' not in session:
        return False
    
    user = User.query.filter_by(userid = session.get('userid')).first()
    if user and user.userid == "d04d23":
        return True
    return False

@app.context_processor
def inject_user():
    # 회원가입 페이지일 때는 항상 로그아웃 상태로 표시
    if request.endpoint == 'register':
        return dict(current_user = None)
    
    # 로그인 상태인 경우에만 current_user 설정
    if 'userid' in session:
        return dict(current_user = session.get('userid'), is_admin = is_admin())
    
    return dict(current_user = None, is_admin = False)

def login_required(f):
   @wraps(f)
   def decorated_function(*args, **kwargs):
       if 'userid' not in session:
           return redirect(url_for('index'))
       return f(*args, **kwargs)
   return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not is_admin():
            flash('관리자 권한이 필요합니다.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

with app.app_context():
    # db.drop_all()
    db.create_all()
   
def validate_userid(userid):
   # 길이 검사
   if len(userid) < 5:
       return False, "아이디는 5자 이상이어야 합니다."
   
   # 공백 검사
   if ' ' in userid:
       return False, "아이디에 공백을 포함할 수 없습니다."  
   
    # 특수문자 검사
   if not re.match("^[a-zA-Z0-9]*$", userid):
       return False, "아이디는 영문자와 숫자만 사용 가능합니다."
   
   # 한글 검사
   if re.search("[ㄱ-ㅎㅏ-ㅣ가-힣]", userid):
       return False, "아이디에 한글을 포함할 수 없습니다."
   
   return True, "유효한 아이디입니다."

@app.route('/')
def index():
    # 통계 데이터 수집
    total_users = User.query.count()
    total_posts = Post.query.count()
    recent_posts = Post.query.order_by(Post.created_at.desc()).limit(5).all()
    
    # 재고 통계
    edl_total = Inventory.query.count()
    edl_normal = Inventory.query.filter_by(status='정상').count()
    smartring_total = SmartringInventory.query.count()
    smartring_normal = SmartringInventory.query.filter_by(status='정상').count()
    
    return render_template('hello.html', 
                         total_users=total_users,
                         total_posts=total_posts,
                         recent_posts=recent_posts,
                         edl_total=edl_total,
                         edl_normal=edl_normal,
                         smartring_total=smartring_total,
                         smartring_normal=smartring_normal)

@app.route('/register', methods=['GET', 'POST'])
def register():
    # GET 요청 처리 - 회원가입 페이지에 접근할 때
    if request.method == 'GET':
        if 'id_checked' in session:
            session.pop('id_checked', None)
        if 'temp_userid' in session:
            session.pop('temp_userid', None)
            
    if request.method == 'POST':
        action = request.form.get('action')
        userid = request.form.get('username')
        
        # 아이디 중복 확인
        if action == 'check':
            valid, message = validate_userid(userid)
            if not valid:
                flash(message, 'error')
                return render_template('register.html', userid = userid)
            
            existing_user = User.query.filter_by(userid = userid).first()
            if existing_user:
                flash('이미 사용 중인 아이디입니다.', 'error')
                # id_checked 세션 변수 설정 안함 (중복된 경우)
                return render_template('register.html', userid = userid)
            
            # 중복 확인 성공 시 id_checked 세션 설정
            session['id_checked'] = True
            session['temp_userid'] = userid  # 임시로 확인된 아이디 저장
            
            flash('사용 가능한 아이디입니다.', 'success')
            return render_template('register.html', userid = userid)
        
        # 회원가입 처리
        elif action == 'register':
            password = request.form.get('password')
            
            # 중복 확인이 되지 않은 경우
            if not session.get('id_checked'):
                flash('아이디 중복 확인이 필요합니다.', 'error')
                return render_template('register.html', userid = userid)
            
            # 중복 확인된 아이디와 현재 입력된 아이디가 다른 경우
            if session.get('temp_userid') != userid:
                flash('중복 확인된 아이디와 다릅니다. 다시 중복 확인해주세요.', 'error')
                session.pop('id_checked', None)
                session.pop('temp_userid', None)
                return render_template('register.html', userid = userid)
            
            # 회원가입 처리
            new_user = User(userid = userid, password = password)
            db.session.add(new_user)
            
            try:
                db.session.commit()
                # 세션 정리
                session.pop('id_checked', None)
                session.pop('temp_userid', None)
                
                flash('회원가입이 완료되었습니다. 로그인해주세요.', 'success')
                return redirect(url_for('index'))
            except Exception as e:
                db.session.rollback()
                flash('회원가입 처리 중 오류가 발생했습니다.', 'error')
                print(f"회원가입 오류: {str(e)}")
                return render_template('register.html', userid = userid)
            
    # GET 요청 처리
    return render_template('register.html')

@app.route('/reset_check', methods=['POST'])
def reset_check():
    if 'id_checked' in session:
        session.pop('id_checked', None)
    if 'temp_userid' in session:
        session.pop('temp_userid', None)
    
    return jsonify({'status': 'success'})

@app.route('/login', methods=['POST'])
def login():
    if request.is_json:
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
    else:
        username = request.form.get('username')
        password = request.form.get('password')
    
    print(f"로그인 시도: 사용자명={username}")
    
    user = User.query.filter_by(userid = username, password = password).first()
    
    if user:
        # 세션에 userid 저장
        session['userid'] = user.userid
        print(f"로그인 성공: 세션 userid={session.get('userid')}")
        
        if 'id_checked' in session:
            session.pop('id_checked', None)
        if 'temp_userid' in session:
            session.pop('temp_userid', None)
            
        print(f"로그인 성공: 세션 userid={session.get('userid')}")
        
        if request.is_json:
            return jsonify({'success': True, 'redirect': url_for('index')})
        else:
            flash(f'{user.userid}님, 환영합니다!', 'success')
            return redirect(url_for('index'))
    else:
        if request.is_json:
            return jsonify({'success': False})
        else:
            flash('아이디 또는 비밀번호가 올바르지 않습니다.', 'error')
            return redirect(url_for('index'))

@app.route('/logout')
def logout():
   session.pop('userid', None)
   flash('로그아웃되었습니다.', 'success')
   return redirect(url_for('index'))

# 파일 확장자 확인 함수
def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/board')
def board_list():
    page = request.args.get('page', 1, type = int)
    per_page = 10
    
    posts = Post.query.order_by(Post.id.desc()).paginate(page = page, per_page = per_page, error_out = False)
    
    return render_template('board_list.html', 
                          posts = posts.items, 
                          current_page = page, 
                          total_pages = posts.pages,
                          current_user = session.get('userid'))
    
@app.route('/board/search')
def board_search():
    keyword = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    if keyword:
        # 제목 또는 내용에 키워드가 포함된 게시글 검색
        posts = Post.query.filter(
            db.or_(
                Post.title.contains(keyword),
                Post.content.contains(keyword),
                Post.author.contains(keyword)
            )
        ).order_by(Post.created_at.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
    else:
        # 검색어가 없으면 전체 목록
        posts = Post.query.order_by(Post.created_at.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
    
    return render_template('board_list.html', 
                          posts=posts.items, 
                          current_page=page, 
                          total_pages=posts.pages,
                          keyword=keyword,
                          current_user=session.get('userid'))
    
@app.route('/board/write', methods=['GET'])
@login_required
def board_write():
    return render_template('board_write.html')

@app.route('/board/write', methods=['POST'])
@login_required
def board_write_post():
    title = request.form.get('title')
    content = request.form.get('content')
    author = session.get('userid')
    
    if not title or not content:
        flash('제목과 내용을 모두 입력해주세요.', 'error')
        return redirect(url_for('board_write'))
    
    # 새 게시글 생성
    new_post = Post(title = title, content = content, author = author)
    
    # 이미지 파일 처리
    if 'image' in request.files:
        file = request.files['image']
        if file and file.filename and allowed_file(file.filename):
            # 파일명 보안 처리 및 중복 방지를 위해 타임스탬프 추가
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            filename = f"{timestamp}_{filename}"
            
            # 업로드 폴더가 없으면 생성
            if not os.path.exists(app.config['UPLOAD_FOLDER']):
                os.makedirs(app.config['UPLOAD_FOLDER'])
                
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            
            # 게시글에 이미지 파일명 저장
            new_post.image_filename = filename
    
    try:
        db.session.add(new_post)
        db.session.commit()
        flash('게시글이 등록되었습니다.', 'success')
    except Exception as e:
        db.session.rollback()
        print(f"게시글 등록 오류: {str(e)}")
        flash('게시글 등록 중 오류가 발생했습니다.', 'error')
        
    return redirect(url_for('board_list'))

@app.route('/board/view/<int:post_id>')
def board_view(post_id):
    # 게시글 조회
    post = Post.query.get_or_404(post_id)
    
    # 조회수 증가
    post.views += 1
    db.session.commit()
    
    # 댓글 조회
    comments = Comment.query.filter_by(post_id = post_id).order_by(Comment.created_at.asc()).all()
    
    return render_template('board_view.html', 
                          post=post, 
                          comments=comments)
    
@app.route('/board/comment/<int:post_id>', methods=['POST'])
@login_required
def comment_write(post_id):
    content = request.form.get('content')
    author = session.get('userid')
    
    if not content:
        flash('댓글 내용을 입력해주세요.', 'error')
        return redirect(url_for('board_view', post_id = post_id))
    
    new_comment = Comment(content = content, author = author, post_id = post_id)
    db.session.add(new_comment)
    db.session.commit()
    
    flash('댓글이 등록되었습니다.', 'success')
    return redirect(url_for('board_view', post_id = post_id))

@app.route('/board/edit/<int:post_id>', methods=['GET'])
@login_required
def board_edit(post_id):
    post = Post.query.get_or_404(post_id)
    
    if post.author != session.get('userid'):
        flash('본인이 작성한 글만 수정할 수 있습니다.', 'error')
        return redirect(url_for('board_view', post_id = post_id))
    
    return render_template('board_edit.html', post = post)

# 게시글 수정 처리
@app.route('/board/edit/<int:post_id>', methods=['POST'])
@login_required
def board_edit_post(post_id):
    post = Post.query.get_or_404(post_id)
    
    if post.author != session.get('userid'):
        flash('수정 권한이 없습니다.', 'error')
        return redirect(url_for('board_list'))
    
    post.title = request.form.get('title')
    post.content = request.form.get('content')
    post.updated_at = datetime.now()
    
    # 이미지 파일 처리
    if 'image' in request.files:
        file = request.files['image']
        if file and file.filename and allowed_file(file.filename):
            # 기존 이미지가 있으면 삭제
            if post.image_filename:
                old_file_path = os.path.join(app.config['UPLOAD_FOLDER'], post.image_filename)
                if os.path.exists(old_file_path):
                    os.remove(old_file_path)
            
            # 새 이미지 저장
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            filename = f"{timestamp}_{filename}"
            
            if not os.path.exists(app.config['UPLOAD_FOLDER']):
                os.makedirs(app.config['UPLOAD_FOLDER'])
                
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            
            # 게시글에 이미지 파일명 업데이트
            post.image_filename = filename
    
    # 이미지 삭제 옵션 처리
    if request.form.get('remove_image') == 'yes' and post.image_filename:
        old_file_path = os.path.join(app.config['UPLOAD_FOLDER'], post.image_filename)
        if os.path.exists(old_file_path):
            os.remove(old_file_path)
        post.image_filename = None
    
    try:
        db.session.commit()
        flash('게시글이 수정되었습니다.', 'success')
    except Exception as e:
        db.session.rollback()
        print(f"게시글 수정 오류: {str(e)}")
        flash('게시글 수정 중 오류가 발생했습니다.', 'error')
    
    return redirect(url_for('board_view', post_id = post_id))

# 게시글 삭제
@app.route('/board/delete/<int:post_id>')
@login_required
def board_delete(post_id):
    post = Post.query.get_or_404(post_id)
    
    if post.author != session.get('userid'):
        flash('삭제 권한이 없습니다.', 'error')
        return redirect(url_for('board_list'))
    
    db.session.delete(post)
    db.session.commit()
    
    flash('게시글이 삭제되었습니다.', 'success')
    return redirect(url_for('board_list'))

# EDL-Doctor 재고 관리 페이지
@app.route('/edl-doctor')
def edl_doctor():
    # 재고 목록 조회
    inventories = Inventory.query.order_by(Inventory.created_at.desc()).all()
    return render_template('edl_doctor.html', inventories=inventories)

@app.route('/edl-doctor/add', methods=['POST'])
@login_required
def add_inventory():
    # 추가 로그인 검사
    if not session.get('userid'):
        flash('재고를 추가하려면 로그인이 필요합니다.', 'error')
        return redirect(url_for('edl_doctor'))
    
    mac_address = request.form.get('mac_address')
    serial_number = request.form.get('serial_number')
    manufacture_date = request.form.get('manufacture_date')
    status = request.form.get('status', '정상')
    location = request.form.get('location')
    note = request.form.get('note')
    
    # 입력값 검증
    if not mac_address or not serial_number or not manufacture_date or not location:
        flash('MAC 주소, 제조번호, 등록일자, 보관위치는 필수 입력 항목입니다.', 'error')
        return redirect(url_for('edl_doctor'))
    
    # MAC 주소 형식 검증 (AA:BB:CC:DD:EE:FF)
    mac_pattern = re.compile(r'^([0-9A-Za-z]{2}[:-]){5}([0-9A-Za-z]{2})$')
    if not mac_pattern.match(mac_address):
        flash('올바른 MAC 주소 형식이 아닙니다. (형식: AA:BB:CC:DD:EE:FF)', 'error')
        return redirect(url_for('edl_doctor'))
    
    # 중복 검사
    exist_mac = Inventory.query.filter_by(mac_address=mac_address).first()
    if exist_mac:
        flash('이미 등록된 MAC 주소입니다.', 'error')
        return redirect(url_for('edl_doctor'))
    
    exist_serial = Inventory.query.filter_by(serial_number=serial_number).first()
    if exist_serial:
        flash('이미 등록된 제조번호입니다.', 'error')
        return redirect(url_for('edl_doctor'))
    
    # 날짜 변환
    try:
        mfg_date = datetime.strptime(manufacture_date, '%Y-%m-%d').date()
    except ValueError:
        flash('올바른 날짜 형식이 아닙니다.', 'error')
        return redirect(url_for('edl_doctor'))
    
    # 재고 추가
    new_inventory = Inventory(
        mac_address=mac_address,
        serial_number=serial_number,
        manufacture_date=mfg_date,
        status=status,
        location=location,
        note=note,
        created_by=session.get('userid')
    )
    
    try:
        db.session.add(new_inventory)
        db.session.commit()
        flash('재고가 성공적으로 추가되었습니다.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'재고 추가 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('edl_doctor'))

@app.route('/edl-doctor/delete/<int:inventory_id>')
@login_required
@admin_required
def delete_inventory(inventory_id):
    # 추가 로그인 검사
    if not session.get('userid'):
        flash('재고를 삭제하려면 로그인이 필요합니다.', 'error')
        return redirect(url_for('edl_doctor'))
    
    if not is_admin():
        flash('재고 삭제 권한이 없습니다.', 'error')
        return redirect(url_for('edl_doctor'))
    
    inventory = Inventory.query.get_or_404(inventory_id)
    
    try:
        db.session.delete(inventory)
        db.session.commit()
        flash('재고가 성공적으로 삭제되었습니다.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'재고 삭제 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('edl_doctor'))

# 재고 상태 업데이트
@app.route('/edl-doctor/update-status/<int:inventory_id>', methods=['POST'])
@login_required
def update_inventory_status(inventory_id):
    # 추가 로그인 검사
    if not session.get('userid'):
        flash('재고 상태를 수정하려면 로그인이 필요합니다.', 'error')
        return redirect(url_for('edl_doctor'))
    
    inventory = Inventory.query.get_or_404(inventory_id)
    mac_address = request.form.get('mac_address')
    serial_number = request.form.get('serial_number')
    new_status = request.form.get('status')
    location = request.form.get('location')
    note = request.form.get('note')
    
    if mac_address != inventory.mac_address:
        exist_mac = Inventory.query.filter(Inventory.mac_address == mac_address, 
                                          Inventory.id != inventory_id).first()
        if exist_mac:
            flash('이미 등록된 MAC 주소입니다.', 'error')
            return redirect(url_for('edl_doctor'))
    
    if serial_number != inventory.serial_number:
        exist_serial = Inventory.query.filter(Inventory.serial_number == serial_number, 
                                             Inventory.id != inventory_id).first()
        if exist_serial:
            flash('이미 등록된 제조번호입니다.', 'error')
            return redirect(url_for('edl_doctor'))
    
    # 변경 사항이 있는지 확인
    has_changes = False
    
    if mac_address and mac_address != inventory.mac_address:
        inventory.mac_address = mac_address
        has_changes = True
    
    if serial_number and serial_number != inventory.serial_number:
        inventory.serial_number = serial_number
        has_changes = True
    
    if new_status and new_status != inventory.status:
        inventory.status = new_status
        has_changes = True
    
    if location != inventory.location:
        inventory.location = location
        has_changes = True
    
    if note is not None and note != inventory.note:
        inventory.note = note
        has_changes = True
    
    if has_changes:
        inventory.updated_at = datetime.now()
    
    try:
        db.session.commit()
        flash('재고 상태가 업데이트되었습니다.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'상태 업데이트 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('edl_doctor'))

# EDL-Doctor Excel 내보내기
@app.route('/edl-doctor/export')
@login_required
def export_edl_inventory():
    inventories = Inventory.query.order_by(Inventory.created_at.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "EDL-Doctor 재고"
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # 헤더 작성
    headers = ['번호', 'MAC 주소', '제조번호', '등록일자', '상태', '보관위치', '수정일자', '등록자', '비고']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 데이터 작성
    for idx, inventory in enumerate(inventories, 2):
        ws.cell(row=idx, column=1, value=idx-1)
        ws.cell(row=idx, column=2, value=inventory.mac_address)
        ws.cell(row=idx, column=3, value=inventory.serial_number)
        ws.cell(row=idx, column=4, value=inventory.manufacture_date.strftime('%Y-%m-%d'))
        ws.cell(row=idx, column=5, value=inventory.status)
        ws.cell(row=idx, column=6, value=inventory.location or '-')
        ws.cell(row=idx, column=7, value=inventory.updated_at.strftime('%Y-%m-%d'))
        ws.cell(row=idx, column=8, value=inventory.created_by or '-')
        ws.cell(row=idx, column=9, value=inventory.note or '-')
    
    # 열 너비 자동 조정
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 파일 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'EDL_Doctor_재고_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# EDL-Doctor Excel 가져오기
@app.route('/edl-doctor/import', methods=['POST'])
@login_required
def import_edl_inventory():
    if 'file' not in request.files:
        flash('파일이 선택되지 않았습니다.', 'error')
        return redirect(url_for('edl_doctor'))
    
    file = request.files['file']
    if file.filename == '':
        flash('파일이 선택되지 않았습니다.', 'error')
        return redirect(url_for('edl_doctor'))
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        flash('Excel 파일만 업로드 가능합니다.', 'error')
        return redirect(url_for('edl_doctor'))
    
    try:
        # Excel 파일 읽기
        try:
            df = pd.read_excel(file, engine='openpyxl', skiprows=0)
        except TypeError as e:
            if 'name should be' in str(e) or '_NamedCellStyle' in str(e):
                file.seek(0)
                df = read_excel_fallback(file)
            else:
                raise e
        
        # NaN 값을 빈 문자열로 변환
        df = df.fillna('')
        
        # 예시 데이터 행 제거 (2번째 행)
        if len(df) > 0:
            first_row = df.iloc[0]
            if 'AA:BB:CC:DD:EE:FF' in str(first_row.get('MAC 주소', '')):
                df = df.iloc[1:]  # 예시 데이터 행 제거
        
        # 설명 행 제거 (※ 작성 안내 이후 행들)
        df = df[df['MAC 주소'].str.contains('※|작성|안내|필수', case=False, na=False) == False]
        df = df[df['MAC 주소'].str.strip() != '']  # 빈 행 제거
        
        # 필수 컬럼 확인
        required_columns = ['MAC 주소', '제조번호', '등록일자', '보관위치']
        if not all(col in df.columns for col in required_columns):
            flash('Excel 파일 형식이 올바르지 않습니다. 필수 컬럼: MAC 주소, 제조번호, 등록일자, 보관위치', 'error')
            return redirect(url_for('edl_doctor'))
        
        success_count = 0
        error_count = 0
        error_messages = []
        
        for index, row in df.iterrows():
            try:
                # 빈 행 건너뛰기
                if not row['MAC 주소'] or str(row['MAC 주소']).strip() == '':
                    continue
                if not row['등록일자'] or str(row['등록일자']).strip() == '':
                    continue
                
                # 중복 확인
                mac_address = str(row['MAC 주소']).strip()
                serial_number = str(row['제조번호']).strip()
                
                existing_mac = Inventory.query.filter_by(mac_address=mac_address).first()
                existing_serial = Inventory.query.filter_by(serial_number=serial_number).first()
                
                if existing_mac or existing_serial:
                    error_count += 1
                    error_messages.append(f"행 {index+2}: 중복된 MAC 주소 또는 제조번호")
                    continue
                
                # 날짜 변환
                manufacture_date = row['등록일자']
                if isinstance(manufacture_date, str):
                    if not manufacture_date.strip():  # 빈 문자열 체크
                        continue  # 빈 행 건너뛰기
                    try:
                        mfg_date = datetime.strptime(manufacture_date, '%Y-%m-%d').date()
                    except ValueError:
                        # Excel 날짜 시리얼 번호 처리
                        serial = int(float(manufacture_date))
                        base_date = datetime(1899, 12, 30)
                        mfg_date = (base_date + timedelta(days=serial)).date()
                elif isinstance(manufacture_date, pd.Timestamp):
                    mfg_date = manufacture_date.date()
                elif isinstance(manufacture_date, (int, float)):
                    # Excel 날짜 시리얼 번호
                    base_date = datetime(1899, 12, 30)
                    mfg_date = (base_date + timedelta(days=int(manufacture_date))).date()
                else:
                    mfg_date = manufacture_date
                
                # 새 재고 추가
                new_inventory = Inventory(
                    mac_address=mac_address,
                    serial_number=serial_number,
                    manufacture_date=mfg_date,
                    status=str(row.get('상태', '정상')).strip() or '정상',
                    location=str(row['보관위치']).strip(),
                    note=str(row.get('비고', '')).strip() if row.get('비고') and str(row.get('비고')).strip() else None,
                    created_by=session.get('userid')
                )
                
                db.session.add(new_inventory)
                success_count += 1
                
            except Exception as e:
                error_count += 1
                error_messages.append(f"행 {index+2}: {str(e)}")
                continue
        
        db.session.commit()
        
        # 결과 메시지
        if success_count > 0:
            flash(f'{success_count}개 항목이 성공적으로 가져왔습니다.', 'success')
        if error_count > 0:
            flash(f'{error_count}개 항목을 가져오는데 실패했습니다.', 'error')
            for msg in error_messages[:5]:  # 최대 5개까지만 표시
                flash(msg, 'error')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Excel 파일 처리 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('edl_doctor'))

# EDL-Doctor 템플릿 다운로드
@app.route('/edl-doctor/template')
def download_edl_template():
    # 새로운 워크북 생성
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거
    ws = wb.create_sheet("EDL-Doctor 재고")
    
    # 헤더 작성 (스타일 없이)
    headers = ['MAC 주소', '제조번호', '등록일자', '상태', '보관위치', '비고']
    ws.append(headers)
    
    # 헤더에만 간단한 스타일 적용
    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # 열 너비 설정
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30
    
    # 파일 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = 'EDL_Doctor_가져오기_템플릿.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# SMARTRING 템플릿 다운로드
@app.route('/smartring/template')
def download_smartring_template():
    # 새로운 워크북 생성
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거
    ws = wb.create_sheet("SMARTRING 재고")
    
    # 헤더 작성 (스타일 없이)
    headers = ['MAC 주소', '제조번호', '등록일자', '상태', '링 사이즈', '색상', '보관위치', '비고']
    ws.append(headers)
    
    # 헤더에만 간단한 스타일 적용
    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # 열 너비 설정
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 30
    
    # 파일 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = 'SMARTRING_가져오기_템플릿.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# SMARTRING 재고 관리 페이지
@app.route('/smartring')
def smartring():
    inventories = SmartringInventory.query.order_by(SmartringInventory.created_at.desc()).all()
    
    # 통계 데이터 계산
    total_count = len(inventories)
    normal_count = sum(1 for inv in inventories if inv.status == '정상')
    repair_count = sum(1 for inv in inventories if inv.status == '수리중')
    defect_count = sum(1 for inv in inventories if inv.status == '불량')
    
    # 링 사이즈별 통계
    size_stats = {}
    for inv in inventories:
        if inv.ring_size:
            size_stats[inv.ring_size] = size_stats.get(inv.ring_size, 0) + 1
    
    # 색상별 통계
    color_stats = {}
    for inv in inventories:
        if inv.color:
            color_stats[inv.color] = color_stats.get(inv.color, 0) + 1
    
    stats = {
        'total': total_count,
        'normal': normal_count,
        'repair': repair_count,
        'defect': defect_count,
        'size_stats': size_stats,
        'color_stats': color_stats
    }
    
    return render_template('Smartring.html', inventories=inventories, stats=stats)

@app.route('/smartring/add', methods=['POST'])
@login_required
def add_smartring():
    if not session.get('userid'):
        flash('재고를 추가하려면 로그인이 필요합니다.', 'error')
        return redirect(url_for('smartring'))
    
    mac_address = request.form.get('mac_address')
    serial_number = request.form.get('serial_number')
    manufacture_date = request.form.get('manufacture_date')
    status = request.form.get('status', '정상')
    location = request.form.get('location')
    ring_size = request.form.get('ring_size')
    color = request.form.get('color')  # 색상 추가
    note = request.form.get('note')
    
    if not mac_address or not serial_number or not manufacture_date or not location:
        flash('MAC 주소, 제조번호, 등록일자, 보관위치는 필수 입력 항목입니다.', 'error')
        return redirect(url_for('smartring'))
    
    mac_pattern = re.compile(r'^([0-9A-Za-z]{2}[:-]){5}([0-9A-Za-z]{2})$')
    if not mac_pattern.match(mac_address):
        flash('올바른 MAC 주소 형식이 아닙니다. (형식: AA:BB:CC:DD:EE:FF)', 'error')
        return redirect(url_for('smartring'))
    
    exist_mac = SmartringInventory.query.filter_by(mac_address=mac_address).first()
    if exist_mac:
        flash('이미 등록된 MAC 주소입니다.', 'error')
        return redirect(url_for('smartring'))
    
    exist_serial = SmartringInventory.query.filter_by(serial_number=serial_number).first()
    if exist_serial:
        flash('이미 등록된 제조번호입니다.', 'error')
        return redirect(url_for('smartring'))
    
    try:
        mfg_date = datetime.strptime(manufacture_date, '%Y-%m-%d').date()
    except ValueError:
        flash('올바른 날짜 형식이 아닙니다.', 'error')
        return redirect(url_for('smartring'))
    
    new_inventory = SmartringInventory(
        mac_address=mac_address,
        serial_number=serial_number,
        manufacture_date=mfg_date,
        status=status,
        location=location,
        ring_size=ring_size,
        color=color,
        note=note,
        created_by=session.get('userid')
    )
    
    try:
        db.session.add(new_inventory)
        db.session.commit()
        flash('재고가 성공적으로 추가되었습니다.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'재고 추가 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('smartring'))

@app.route('/smartring/delete/<int:inventory_id>')
@login_required
@admin_required
def delete_smartring(inventory_id):
    if not session.get('userid'):
        flash('재고를 삭제하려면 로그인이 필요합니다.', 'error')
        return redirect(url_for('smartring'))
    
    if not is_admin():
        flash('재고 삭제 권한이 없습니다.', 'error')
        return redirect(url_for('smartring'))
    
    inventory = SmartringInventory.query.get_or_404(inventory_id)
    
    try:
        db.session.delete(inventory)
        db.session.commit()
        flash('재고가 성공적으로 삭제되었습니다.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'재고 삭제 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('smartring'))

# SMARTRING 재고 상태 업데이트
@app.route('/smartring/update-status/<int:inventory_id>', methods=['POST'])
@login_required
def update_smartring_status(inventory_id):
    if not session.get('userid'):
        flash('재고 상태를 수정하려면 로그인이 필요합니다.', 'error')
        return redirect(url_for('smartring'))
    
    inventory = SmartringInventory.query.get_or_404(inventory_id)
    mac_address = request.form.get('mac_address')
    serial_number = request.form.get('serial_number')
    new_status = request.form.get('status')
    location = request.form.get('location')
    ring_size = request.form.get('ring_size')
    color = request.form.get('color')
    note = request.form.get('note')
    
    if mac_address != inventory.mac_address:
        exist_mac = SmartringInventory.query.filter(SmartringInventory.mac_address == mac_address, 
                                          SmartringInventory.id != inventory_id).first()
        if exist_mac:
            flash('이미 등록된 MAC 주소입니다.', 'error')
            return redirect(url_for('smartring'))
    
    if serial_number != inventory.serial_number:
        exist_serial = SmartringInventory.query.filter(SmartringInventory.serial_number == serial_number, 
                                             SmartringInventory.id != inventory_id).first()
        if exist_serial:
            flash('이미 등록된 제조번호입니다.', 'error')
            return redirect(url_for('smartring'))
    
    has_changes = False
    
    if mac_address and mac_address != inventory.mac_address:
        inventory.mac_address = mac_address
        has_changes = True
    
    if serial_number and serial_number != inventory.serial_number:
        inventory.serial_number = serial_number
        has_changes = True
    
    if new_status and new_status != inventory.status:
        inventory.status = new_status
        has_changes = True
    
    if location != inventory.location:
        inventory.location = location
        has_changes = True
    
    if ring_size != inventory.ring_size:
        inventory.ring_size = ring_size
        has_changes = True
    
    if color != inventory.color:
        inventory.color = color
        has_changes = True
    
    if note is not None and note != inventory.note:
        inventory.note = note
        has_changes = True
    
    if has_changes:
        inventory.updated_at = datetime.now()
    
    try:
        db.session.commit()
        flash('재고 상태가 업데이트되었습니다.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'상태 업데이트 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('smartring'))

# SMARTRING Excel 내보내기
@app.route('/smartring/export')
@login_required
def export_smartring_inventory():
    inventories = SmartringInventory.query.order_by(SmartringInventory.created_at.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "SMARTRING 재고"
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # 헤더 작성
    headers = ['번호', 'MAC 주소', '제조번호', '등록일자', '상태', '링 사이즈', '색상', '보관위치', '수정일자', '등록자', '비고']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 데이터 작성
    for idx, inventory in enumerate(inventories, 2):
        ws.cell(row=idx, column=1, value=idx-1)
        ws.cell(row=idx, column=2, value=inventory.mac_address)
        ws.cell(row=idx, column=3, value=inventory.serial_number)
        ws.cell(row=idx, column=4, value=inventory.manufacture_date.strftime('%Y-%m-%d'))
        ws.cell(row=idx, column=5, value=inventory.status)
        ws.cell(row=idx, column=6, value=inventory.ring_size or '-')
        ws.cell(row=idx, column=7, value=inventory.color or '-')
        ws.cell(row=idx, column=8, value=inventory.location or '-')
        ws.cell(row=idx, column=9, value=inventory.updated_at.strftime('%Y-%m-%d'))
        ws.cell(row=idx, column=10, value=inventory.created_by or '-')
        ws.cell(row=idx, column=11, value=inventory.note or '-')
    
    # 열 너비 자동 조정
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 파일 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'SMARTRING_재고_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# SMARTRING Excel 가져오기
@app.route('/smartring/import', methods=['POST'])
@login_required
def import_smartring_inventory():
    if 'file' not in request.files:
        flash('파일이 선택되지 않았습니다.', 'error')
        return redirect(url_for('smartring'))
    
    file = request.files['file']
    if file.filename == '':
        flash('파일이 선택되지 않았습니다.', 'error')
        return redirect(url_for('smartring'))
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        flash('Excel 파일만 업로드 가능합니다.', 'error')
        return redirect(url_for('smartring'))
    
    try:
        # Excel 파일 읽기
        try:
            df = pd.read_excel(file, engine='openpyxl', skiprows=0)
        except TypeError as e:
            if 'name should be' in str(e) or '_NamedCellStyle' in str(e):
                file.seek(0)
                df = read_excel_fallback(file)
            else:
                raise e
        
        # NaN 값을 빈 문자열로 변환
        df = df.fillna('')
        
        # 예시 데이터 행 제거 (2번째 행)
        if len(df) > 0:
            first_row = df.iloc[0]
            if 'AA:BB:CC:DD:EE:FF' in str(first_row.get('MAC 주소', '')):
                df = df.iloc[1:]  # 예시 데이터 행 제거
        
        # 설명 행 제거 (※ 작성 안내 이후 행들)
        df = df[df['MAC 주소'].str.contains('※|작성|안내|필수', case=False, na=False) == False]
        df = df[df['MAC 주소'].str.strip() != '']  # 빈 행 제거
        
        # 필수 컬럼 확인
        required_columns = ['MAC 주소', '제조번호', '등록일자', '보관위치']
        if not all(col in df.columns for col in required_columns):
            flash('Excel 파일 형식이 올바르지 않습니다. 필수 컬럼: MAC 주소, 제조번호, 등록일자, 보관위치', 'error')
            return redirect(url_for('smartring'))
        
        success_count = 0
        error_count = 0
        error_messages = []
        
        for index, row in df.iterrows():
            try:
                # 빈 행 건너뛰기
                if not row['MAC 주소'] or str(row['MAC 주소']).strip() == '':
                    continue
                if not row['등록일자'] or str(row['등록일자']).strip() == '':
                    continue
                
                # 중복 확인
                mac_address = str(row['MAC 주소']).strip()
                serial_number = str(row['제조번호']).strip()
                
                existing_mac = SmartringInventory.query.filter_by(mac_address=mac_address).first()
                existing_serial = SmartringInventory.query.filter_by(serial_number=serial_number).first()
                
                if existing_mac or existing_serial:
                    error_count += 1
                    error_messages.append(f"행 {index+2}: 중복된 MAC 주소 또는 제조번호")
                    continue
                
                # 날짜 변환
                manufacture_date = row['등록일자']
                if isinstance(manufacture_date, str):
                    if not manufacture_date.strip():  # 빈 문자열 체크
                        continue  # 빈 행 건너뛰기
                    try:
                        mfg_date = datetime.strptime(manufacture_date, '%Y-%m-%d').date()
                    except ValueError:
                        # Excel 날짜 시리얼 번호 처리
                        serial = int(float(manufacture_date))
                        base_date = datetime(1899, 12, 30)
                        mfg_date = (base_date + timedelta(days=serial)).date()
                elif isinstance(manufacture_date, pd.Timestamp):
                    mfg_date = manufacture_date.date()
                elif isinstance(manufacture_date, (int, float)):
                    # Excel 날짜 시리얼 번호
                    base_date = datetime(1899, 12, 30)
                    mfg_date = (base_date + timedelta(days=int(manufacture_date))).date()
                else:
                    mfg_date = manufacture_date
                
                # 링 사이즈와 색상 처리
                ring_size = str(row.get('링 사이즈', '')).strip() if row.get('링 사이즈') else None
                color = str(row.get('색상', '')).strip() if row.get('색상') else None
                
                # 새 재고 추가
                new_inventory = SmartringInventory(
                    mac_address=mac_address,
                    serial_number=serial_number,
                    manufacture_date=mfg_date,
                    status=str(row.get('상태', '정상')).strip() or '정상',
                    ring_size=ring_size if ring_size else None,
                    color=color if color else None,
                    location=str(row['보관위치']).strip(),
                    note=str(row.get('비고', '')).strip() if row.get('비고') and str(row.get('비고')).strip() else None,
                    created_by=session.get('userid')
                )
                
                db.session.add(new_inventory)
                success_count += 1
                
            except Exception as e:
                error_count += 1
                error_messages.append(f"행 {index+2}: {str(e)}")
                continue
        
        db.session.commit()
        
        # 결과 메시지
        if success_count > 0:
            flash(f'{success_count}개 항목이 성공적으로 가져왔습니다.', 'success')
        if error_count > 0:
            flash(f'{error_count}개 항목을 가져오는데 실패했습니다.', 'error')
            for msg in error_messages[:5]:  # 최대 5개까지만 표시
                flash(msg, 'error')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Excel 파일 처리 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('smartring'))

# 나의 정보 페이지
@app.route('/my_info')
@login_required
def my_info():
    user = User.query.filter_by(userid=session.get('userid')).first()
    if not user:
        flash('사용자 정보를 찾을 수 없습니다.', 'error')
        return redirect(url_for('index'))
    
    return render_template('my_info.html', user=user)

# 비밀번호 변경
@app.route('/change_password', methods=['POST'])
@login_required
def change_password():
    current_password = request.form.get('current_password')
    new_password = request.form.get('new_password')
    confirm_password = request.form.get('confirm_password')
    
    # 현재 사용자 확인
    user = User.query.filter_by(userid=session.get('userid')).first()
    if not user:
        flash('사용자 정보를 찾을 수 없습니다.', 'error')
        return redirect(url_for('index'))
    
    # 현재 비밀번호 확인
    if user.password != current_password:
        flash('현재 비밀번호가 올바르지 않습니다.', 'error')
        return redirect(url_for('my_info'))
    
    # 새 비밀번호 검증
    if new_password != confirm_password:
        flash('새 비밀번호가 일치하지 않습니다.', 'error')
        return redirect(url_for('my_info'))
    
    if len(new_password) < 4:
        flash('비밀번호는 최소 4자 이상이어야 합니다.', 'error')
        return redirect(url_for('my_info'))
    
    # 비밀번호 변경
    try:
        user.password = new_password
        db.session.commit()
        flash('비밀번호가 성공적으로 변경되었습니다.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'비밀번호 변경 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('my_info'))

# 회원 탈퇴
@app.route('/delete_account', methods=['POST'])
@login_required
def delete_account():
    password = request.form.get('password')
    
    # 현재 사용자 확인
    user = User.query.filter_by(userid=session.get('userid')).first()
    if not user:
        flash('사용자 정보를 찾을 수 없습니다.', 'error')
        return redirect(url_for('index'))
    
    # 비밀번호 확인
    if user.password != password:
        flash('비밀번호가 올바르지 않습니다.', 'error')
        return redirect(url_for('my_info'))
    
    # 사용자와 관련된 게시글, 댓글, 재고 정보 삭제
    try:
        # 사용자가 작성한 다른 게시글의 댓글 먼저 삭제
        Comment.query.filter_by(author=user.userid).delete()

        # 사용자가 작성한 게시글 삭제
        posts = Post.query.filter_by(author=user.userid).all()
        for post in posts:
            db.session.delete(post)
        
        # 사용자가 등록한 재고 정보는 created_by만 NULL로 변경
        inventories = Inventory.query.filter_by(created_by=user.userid).all()
        for inventory in inventories:
            inventory.created_by = None
        
        # 사용자 계정 삭제
        db.session.delete(user)
        db.session.commit()
        
        # 세션 제거
        session.pop('userid', None)
        
        flash('회원 탈퇴가 완료되었습니다.', 'success')
        return redirect(url_for('index'))
    except Exception as e:
        db.session.rollback()
        flash(f'회원 탈퇴 처리 중 오류가 발생했습니다: {str(e)}', 'error')
        return redirect(url_for('my_info'))

if __name__ == '__main__':
   app.run(host='127.0.0.1', port = 5000, debug = True)